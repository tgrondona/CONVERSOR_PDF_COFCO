"""
pdf_a_excel.py
──────────────
Convierte un estado de cuenta del corredor (PDF) a un Excel tabulado y prolijo.

Uso:
    python pdf_a_excel.py  <ruta_pdf>  [ruta_salida.xlsx]

Ejemplo:
    python pdf_a_excel.py COFCO.pdf COFCO_movimientos.xlsx
"""

import sys
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from parse_broker_pdf import parse_broker_pdf


# ═══════════════════════════════════════════════════════════════
# ESTILOS
# ═══════════════════════════════════════════════════════════════

C_HEADER = 'FF1F3864'  # azul oscuro
C_SUBHDR = 'FF2E75B6'  # azul medio
C_SECUSD = 'FF203864'  # sección USD
C_SECARS = 'FF375623'  # sección ARS verde oscuro
C_SECDC  = 'FF7B3F00'  # sección dif cambio marrón
C_ALT    = 'FFD9E1F2'  # fila par azul
C_ALT2   = 'FFE2EFDA'  # fila par verde (ARS)
C_ALT3   = 'FFFFF2CC'  # fila par amarillo (DC)
C_WHITE  = 'FFFFFFFF'
C_TOTROW = 'FFBDD7EE'  # total row azul claro

_thin = Side(style='thin', color='FFB0B0B0')
_thick = Side(style='medium', color='FF666666')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_TOP = Border(left=_thin, right=_thin, top=_thick, bottom=_thin)

FMT_ARS  = '#,##0.00;[RED](#,##0.00);"-"'
FMT_USD  = '#,##0.00;[RED](#,##0.00);"-"'
FMT_DATE = 'DD/MM/YYYY'


def fill(c): return PatternFill('solid', start_color=c)

def sh(cell, bg, txt='FFFFFFFF', bold=True, sz=10, align='center'):
    cell.fill = fill(bg)
    cell.font = Font(name='Calibri', bold=bold, size=sz, color=txt)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = BORDER

def sc(cell, fmt=None, bold=False, align='left', bg=C_WHITE, txt='FF000000', top_border=False):
    cell.font = Font(name='Calibri', bold=bold, size=10, color=txt)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = BORDER_TOP if top_border else BORDER
    cell.fill = fill(bg)
    if fmt:
        cell.number_format = fmt


# ═══════════════════════════════════════════════════════════════
# ETIQUETAS LEGIBLES
# ═══════════════════════════════════════════════════════════════

SECCION_LABELS = {
    'USD_vencido':  'CUENTA EN DÓLARES — Saldo Vencido',
    'USD_avencer':  'CUENTA EN DÓLARES — Saldo a Vencer',
    'ARS_vencido':  'CUENTA EN PESOS — Saldo Vencido',
    'ARS_avencer':  'CUENTA EN PESOS — Saldo a Vencer',
    'DIF_CAMBIO':   'DIFERENCIA DE CAMBIO',
}

SECCION_COLORS = {
    'USD_vencido':  (C_SECUSD, C_ALT),
    'USD_avencer':  (C_SECUSD, C_ALT),
    'ARS_vencido':  (C_SECARS, C_ALT2),
    'ARS_avencer':  (C_SECARS, C_ALT2),
    'DIF_CAMBIO':   (C_SECDC,  C_ALT3),
}


# ═══════════════════════════════════════════════════════════════
# SHEET: TODOS LOS MOVIMIENTOS (hoja plana)
# ═══════════════════════════════════════════════════════════════

HEADERS = [
    ('Sección',         16),
    ('Fecha',           12),
    ('Vto.',            12),
    ('Tipo Doc.',       22),
    ('Comprobante',     22),
    ('Comp. Original',  18),
    ('Concepto',        28),
    ('Contrato',        16),
    ('Moneda',           8),
    ('Importe',         16),
    ('Saldo',           16),
]

def sheet_movimientos(wb, df, pdf_name, pdf_totals):
    ws = wb.create_sheet('MOVIMIENTOS')

    # Título
    ncols = len(HEADERS)
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    c = ws.cell(1, 1, f'Estado de Cuenta — {pdf_name}')
    sh(c, C_HEADER, sz=13)
    ws.row_dimensions[1].height = 28

    # Headers
    for ci, (h, _) in enumerate(HEADERS, 1):
        sh(ws.cell(2, ci, h), C_SUBHDR)
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = 'A3'

    row_num = 3
    order = ['USD_vencido', 'USD_avencer', 'ARS_vencido', 'ARS_avencer', 'DIF_CAMBIO']

    for sec in order:
        sub = df[df['seccion'] == sec].copy()
        if sub.empty:
            continue

        hdr_color, alt_color = SECCION_COLORS[sec]

        # Separador de sección
        ws.merge_cells(f'A{row_num}:{get_column_letter(ncols)}{row_num}')
        sh(ws.cell(row_num, 1, SECCION_LABELS[sec]), hdr_color, sz=11)
        ws.row_dimensions[row_num].height = 20
        row_num += 1

        # Filas de datos
        for i, (_, r) in enumerate(sub.iterrows()):
            bg = alt_color if i % 2 == 0 else C_WHITE
            ws.row_dimensions[row_num].height = 15

            vals = [
                SECCION_LABELS[sec],
                r['fecha'],
                r['vto'],
                r['tipo_doc'],
                r['comprobante'],
                r['comprobante_orig'] or '',
                r['concepto'],
                r['contrato'],
                r['moneda'],
                r['importe'],
                r['saldo'] if abs(r['saldo']) > 0.001 else None,
            ]

            for ci, (v, (_, w)) in enumerate(zip(vals, HEADERS), 1):
                cell = ws.cell(row_num, ci, v if v is not None else '')
                if ci == 2 or ci == 3:      # fechas
                    sc(cell, fmt=FMT_DATE, align='center', bg=bg)
                elif ci == 9:               # moneda
                    sc(cell, align='center', bg=bg)
                elif ci in (10, 11):        # importes
                    mon = r['moneda']
                    fmt = FMT_USD if mon == 'USD' else FMT_ARS
                    sc(cell, fmt=fmt, align='right', bg=bg)
                else:
                    sc(cell, align='left', bg=bg)

            row_num += 1

        # Fila de subtotal de sección
        ws.merge_cells(f'A{row_num}:I{row_num}')
        subtotal_label = f'Subtotal — {SECCION_LABELS[sec]}'
        c_lbl = ws.cell(row_num, 1, subtotal_label)
        sc(c_lbl, bold=True, bg=C_TOTROW, top_border=True)

        mon = 'USD' if sec.startswith('USD') else 'ARS'
        fmt = FMT_USD if mon == 'USD' else FMT_ARS
        c_imp = ws.cell(row_num, 10, sub['importe'].sum())
        sc(c_imp, fmt=fmt, bold=True, align='right', bg=C_TOTROW, top_border=True)
        c_sal = ws.cell(row_num, 11, sub['saldo'].sum() if sub['saldo'].abs().sum() > 0.01 else None)
        sc(c_sal, fmt=fmt, bold=True, align='right', bg=C_TOTROW, top_border=True)

        ws.row_dimensions[row_num].height = 16
        row_num += 2  # línea en blanco

    # Anchos de columna
    for ci, (_, w) in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ═══════════════════════════════════════════════════════════════
# SHEET: UNA HOJA POR SECCIÓN
# ═══════════════════════════════════════════════════════════════

SECTION_SHEETS = [
    ('USD_vencido',  'USD Vencido',
     [('Fecha',          'fecha',         FMT_DATE, 12, 'center'),
      ('Vencimiento',    'vto',           FMT_DATE, 12, 'center'),
      ('Tipo Doc.',      'tipo_doc',      None,     20, 'left'),
      ('Comprobante',    'comprobante',   None,     22, 'left'),
      ('Concepto',       'concepto',      None,     26, 'left'),
      ('Contrato',       'contrato',      None,     16, 'left'),
      ('Moneda',         'moneda',        None,      8, 'center'),
      ('Importe USD',    'importe',       FMT_USD,  16, 'right'),
      ('Saldo USD',      'saldo',         FMT_USD,  14, 'right')]),

    ('USD_avencer',  'USD A Vencer',
     [('Fecha',          'fecha',         FMT_DATE, 12, 'center'),
      ('Vencimiento',    'vto',           FMT_DATE, 12, 'center'),
      ('Tipo Doc.',      'tipo_doc',      None,     20, 'left'),
      ('Comprobante',    'comprobante',   None,     22, 'left'),
      ('Concepto',       'concepto',      None,     26, 'left'),
      ('Contrato',       'contrato',      None,     16, 'left'),
      ('Moneda',         'moneda',        None,      8, 'center'),
      ('Importe USD',    'importe',       FMT_USD,  16, 'right'),
      ('Saldo USD',      'saldo',         FMT_USD,  14, 'right')]),

    ('ARS_vencido',  'ARS Vencido',
     [('Fecha',          'fecha',         FMT_DATE, 12, 'center'),
      ('Vencimiento',    'vto',           FMT_DATE, 12, 'center'),
      ('Tipo Doc.',      'tipo_doc',      None,     22, 'left'),
      ('Comprobante',    'comprobante',   None,     22, 'left'),
      ('Concepto',       'concepto',      None,     28, 'left'),
      ('Contrato',       'contrato',      None,     16, 'left'),
      ('Moneda',         'moneda',        None,      8, 'center'),
      ('Importe ARS',    'importe',       FMT_ARS,  18, 'right'),
      ('Saldo ARS',      'saldo',         FMT_ARS,  16, 'right')]),

    ('ARS_avencer',  'ARS A Vencer',
     [('Fecha',          'fecha',         FMT_DATE, 12, 'center'),
      ('Vencimiento',    'vto',           FMT_DATE, 12, 'center'),
      ('Tipo Doc.',      'tipo_doc',      None,     22, 'left'),
      ('Comprobante',    'comprobante',   None,     22, 'left'),
      ('Concepto',       'concepto',      None,     28, 'left'),
      ('Contrato',       'contrato',      None,     16, 'left'),
      ('Moneda',         'moneda',        None,      8, 'center'),
      ('Importe ARS',    'importe',       FMT_ARS,  18, 'right'),
      ('Saldo ARS',      'saldo',         FMT_ARS,  16, 'right')]),

    ('DIF_CAMBIO',   'Dif. Cambio',
     [('Fecha',          'fecha',         FMT_DATE, 12, 'center'),
      ('Vencimiento',    'vto',           FMT_DATE, 12, 'center'),
      ('Tipo Doc.',      'tipo_doc',      None,     18, 'left'),
      ('Comprobante',    'comprobante',   None,     22, 'left'),
      ('Comp. Original', 'comprobante_orig', None,  22, 'left'),
      ('Contrato',       'contrato',      None,     10, 'left'),
      ('Moneda',         'moneda',        None,      8, 'center'),
      ('Importe ARS',    'importe',       FMT_ARS,  18, 'right'),
      ('Saldo ARS',      'saldo',         FMT_ARS,  16, 'right')]),
]


def sheet_seccion(wb, sec_key, sheet_name, col_specs, df_sec, pdf_name):
    if df_sec.empty:
        return

    ws = wb.create_sheet(sheet_name)
    hdr_color, alt_color = SECCION_COLORS[sec_key]
    ncols = len(col_specs)

    # Título
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    title = f'{SECCION_LABELS[sec_key]}  —  {pdf_name}'
    sh(ws.cell(1, 1, title), hdr_color, sz=12)
    ws.row_dimensions[1].height = 26

    # Cabecera
    for ci, (h, _, _, w, _) in enumerate(col_specs, 1):
        sh(ws.cell(2, ci, h), C_SUBHDR)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 26
    ws.freeze_panes = 'A3'

    # Datos
    for i, (_, row) in enumerate(df_sec.iterrows()):
        rn = 3 + i
        ws.row_dimensions[rn].height = 15
        bg = alt_color if i % 2 == 0 else C_WHITE

        for ci, (_, col, fmt, _, align) in enumerate(col_specs, 1):
            v = row.get(col, '')
            if pd.isna(v) or v == 'nan':
                v = ''
            # Para saldo, ocultar ceros
            if col == 'saldo' and isinstance(v, (int, float)) and abs(v) < 0.001:
                v = None
            cell = ws.cell(rn, ci, v if v is not None else '')
            sc(cell, fmt=fmt, align=align, bg=bg)

    # Total
    rn_tot = 3 + len(df_sec)
    ws.merge_cells(f'A{rn_tot}:{get_column_letter(ncols-2)}{rn_tot}')
    c_lbl = ws.cell(rn_tot, 1, f'TOTAL  ({len(df_sec)} movimientos)')
    sc(c_lbl, bold=True, bg=C_TOTROW, top_border=True)

    # Sumar columnas numéricas
    for ci, (_, col, fmt, _, _) in enumerate(col_specs, 1):
        if fmt in (FMT_ARS, FMT_USD) and col in df_sec.columns:
            total = df_sec[col].sum()
            c = ws.cell(rn_tot, ci, total)
            sc(c, fmt=fmt, bold=True, align='right', bg=C_TOTROW, top_border=True)

    ws.row_dimensions[rn_tot].height = 18


# ═══════════════════════════════════════════════════════════════
# SHEET: RESUMEN / PORTADA
# ═══════════════════════════════════════════════════════════════

def sheet_resumen(wb, df, pdf_totals, pdf_name, pdf_path):
    ws = wb.create_sheet('RESUMEN', 0)  # primera hoja

    ws.merge_cells('A1:E1')
    sh(ws.cell(1, 1, f'ESTADO DE CUENTA — {pdf_name}'), C_HEADER, sz=14)
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:E2')
    from datetime import date
    sh(ws.cell(2, 1, f'Extraído el {date.today().strftime("%d/%m/%Y")}  |  Archivo: {Path(pdf_path).name}'),
       C_SUBHDR, sz=10, bold=False)
    ws.row_dimensions[2].height = 18

    # Tabla resumen por sección
    r = 4
    for ci, h in enumerate(['Sección', 'Movimientos', 'Importe', 'Saldo Pendiente', 'Moneda'], 1):
        sh(ws.cell(r, ci), C_SUBHDR)
    ws.row_dimensions[r].height = 24
    r += 1

    order = ['USD_vencido', 'USD_avencer', 'ARS_vencido', 'ARS_avencer', 'DIF_CAMBIO']
    for i, sec in enumerate(order):
        sub = df[df['seccion'] == sec]
        if sub.empty:
            continue
        bg = SECCION_COLORS[sec][1] if i % 2 == 0 else C_WHITE
        mon = 'USD' if sec.startswith('USD') else 'ARS'
        fmt = FMT_USD if mon == 'USD' else FMT_ARS

        ws.row_dimensions[r].height = 16
        sc(ws.cell(r, 1, SECCION_LABELS[sec]), bold=True, bg=bg)
        sc(ws.cell(r, 2, len(sub)), align='center', bg=bg)
        c_imp = ws.cell(r, 3, sub['importe'].sum())
        sc(c_imp, fmt=fmt, align='right', bg=bg)
        c_sal = ws.cell(r, 4, sub['saldo'].sum())
        sc(c_sal, fmt=fmt, align='right', bg=bg)
        sc(ws.cell(r, 5, mon), align='center', bg=bg)
        r += 1

    # Totales extraídos del PDF
    r += 1
    ws.merge_cells(f'A{r}:E{r}')
    sh(ws.cell(r, 1, 'SALDOS FINALES (del PDF)'), C_SUBHDR)
    ws.row_dimensions[r].height = 20
    r += 1

    for ci, h in enumerate(['Cuenta', 'Saldo PDF'], 1):
        sh(ws.cell(r, ci), C_HEADER)
    ws.row_dimensions[r].height = 22
    r += 1

    saldos = [
        ('Cuenta en Dólares (USD)', pdf_totals.get('total_USD', 0), FMT_USD),
        ('Cuenta en Pesos (ARS)',   pdf_totals.get('total_ARS', 0), FMT_ARS),
    ]
    for i, (label, val, fmt) in enumerate(saldos):
        bg = C_ALT if i % 2 == 0 else C_WHITE
        ws.row_dimensions[r].height = 16
        sc(ws.cell(r, 1, label), bold=True, bg=bg)
        c = ws.cell(r, 2, val or 0)
        sc(c, fmt=fmt, bold=True, align='right', bg=bg)
        r += 1

    # Anchos
    for col, w in zip('ABCDE', [40, 14, 18, 18, 10]):
        ws.column_dimensions[col].width = w


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

def pdf_a_excel(pdf_path, output_path=None):
    pdf_path = Path(pdf_path)
    if output_path is None:
        output_path = pdf_path.parent / f'{pdf_path.stem}_movimientos.xlsx'
    output_path = Path(output_path)

    print(f"Leyendo PDF: {pdf_path.name} ...")
    result = parse_broker_pdf(str(pdf_path))
    df = result['transactions']
    totals = result['totals']

    if df.empty:
        print("⚠ No se encontraron movimientos en el PDF.")
        return

    print(f"  {len(df)} movimientos en {df['seccion'].nunique()} secciones")

    wb = Workbook()
    wb.remove(wb.active)

    # Hoja 1: Resumen
    sheet_resumen(wb, df, totals, pdf_path.stem.upper(), pdf_path)

    # Hoja 2: Todos los movimientos juntos
    sheet_movimientos(wb, df, pdf_path.stem.upper(), totals)

    # Hojas por sección
    for sec_key, sheet_name, col_specs in SECTION_SHEETS:
        df_sec = df[df['seccion'] == sec_key].copy()
        sheet_seccion(wb, sec_key, sheet_name, col_specs, df_sec, pdf_path.stem.upper())

    wb.save(str(output_path))
    print(f"✅ Excel guardado: {output_path.name}")
    return output_path


if __name__ == '__main__':
    args = sys.argv[1:]
    if not args:
        print("Uso: python pdf_a_excel.py <archivo.pdf> [salida.xlsx]")
        sys.exit(1)
    out = pdf_a_excel(args[0], args[1] if len(args) > 1 else None)
