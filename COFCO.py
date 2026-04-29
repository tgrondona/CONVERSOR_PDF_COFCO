import sys
from pathlib import Path
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# IMPORTANTE: este archivo tiene que estar en la misma carpeta
from parse_broker_pdf import parse_broker_pdf


# ═══════════════════════════════════════════════════════════════
# ESTILOS
# ═══════════════════════════════════════════════════════════════

C_HEADER = 'FF1F3864'
C_SUBHDR = 'FF2E75B6'
C_SECUSD = 'FF203864'
C_SECARS = 'FF375623'
C_SECDC  = 'FF7B3F00'
C_ALT    = 'FFD9E1F2'
C_ALT2   = 'FFE2EFDA'
C_ALT3   = 'FFFFF2CC'
C_WHITE  = 'FFFFFFFF'
C_TOTROW = 'FFBDD7EE'

_thin = Side(style='thin', color='FFB0B0B0')
_thick = Side(style='medium', color='FF666666')

BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_TOP = Border(left=_thin, right=_thin, top=_thick, bottom=_thin)

FMT_ARS  = '#,##0.00;[RED](#,##0.00);"-"'
FMT_USD  = '#,##0.00;[RED](#,##0.00);"-"'
FMT_DATE = 'DD/MM/YYYY'


def fill(c): 
    return PatternFill('solid', start_color=c)

def sh(cell, bg, txt='FFFFFFFF', bold=True, sz=10, align='center'):
    cell.fill = fill(bg)
    cell.font = Font(name='Calibri', bold=bold, size=sz, color=txt)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = BORDER

def sc(cell, fmt=None, bold=False, align='left', bg=C_WHITE, txt='FF000000', top_border=False):
    cell.font = Font(name='Calibri', bold=bold, size=10, color=txt)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = BORDER_TOP if top_border else BORDER
    cell.fill = fill(bg)
    if fmt:
        cell.number_format = fmt


# ═══════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════

SECCION_LABELS = {
    'USD_vencido':  'CUENTA EN DÓLARES — Saldo Vencido',
    'USD_avencer':  'CUENTA EN DÓLARES — Saldo a Vencer',
    'ARS_vencido':  'CUENTA EN PESOS — Saldo Vencido',
    'ARS_avencer':  'CUENTA EN PESOS — Saldo a Vencer',
    'DIF_CAMBIO':   'DIFERENCIA DE CAMBIO',
}

SECCION_COLORS = {
    'USD_vencido':  (C_SECUSD, C_ALT),
    'USD_avencer':  (C_SECUSD, C_ALT),
    'ARS_vencido':  (C_SECARS, C_ALT2),
    'ARS_avencer':  (C_SECARS, C_ALT2),
    'DIF_CAMBIO':   (C_SECDC,  C_ALT3),
}


# ═══════════════════════════════════════════════════════════════
# CORE
# ═══════════════════════════════════════════════════════════════

def pdf_a_excel(pdf_path, output_path=None):
    pdf_path = Path(pdf_path)

    if output_path is None:
        output_path = pdf_path.with_suffix('.xlsx')

    print(f"Leyendo PDF: {pdf_path.name}")

    result = parse_broker_pdf(str(pdf_path))
    df = result['transactions']
    totals = result['totals']

    if df.empty:
        raise Exception("No se encontraron movimientos")

    wb = Workbook()
    ws = wb.active
    ws.title = "MOVIMIENTOS"

    headers = list(df.columns)

    # headers
    for col, h in enumerate(headers, 1):
        sh(ws.cell(1, col, h), C_SUBHDR)

    # data
    for i, row in df.iterrows():
        for col, h in enumerate(headers, 1):
            ws.cell(i + 2, col, row[h])

    wb.save(output_path)

    return output_path


# ═══════════════════════════════════════════════════════════════
# UI (selector de archivo)
# ═══════════════════════════════════════════════════════════════

import tkinter as tk
from tkinter import filedialog, messagebox


def ejecutar_app():
    root = tk.Tk()
    root.withdraw()

    pdf_path = filedialog.askopenfilename(
        title="Seleccionar PDF",
        filetypes=[("PDF", "*.pdf")]
    )

    if not pdf_path:
        return

    output_path = filedialog.asksaveasfilename(
        title="Guardar Excel como",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")]
    )

    if not output_path:
        return

    try:
        result = pdf_a_excel(pdf_path, output_path)
        messagebox.showinfo("OK", f"Archivo generado:\n{result}")
    except Exception as e:
        messagebox.showerror("Error", str(e))


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    ejecutar_app()